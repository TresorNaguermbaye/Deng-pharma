# apps/reporting/views.py
from datetime import date
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import Coalesce

from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError

from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.sales.models import Sale
from apps.inventory.models import StockLot
from apps.notifications.models import Notification
from apps.accounts.models import User






def get_pharmacy_info(request):
    """Renvoie les informations de la pharmacie pour l'en-tête des rapports."""
    admin_user = User.objects.filter(role='ADMIN').first()
    logo_url = None
    if admin_user and hasattr(admin_user, 'profile') and admin_user.profile and admin_user.profile.photo:
        logo_url = request.build_absolute_uri(admin_user.profile.photo.url)

    return {
        'logo_url': logo_url,
        'name': 'DENG PHARMA',
        'email': admin_user.email if admin_user else '',
        'phone': '+235 66 00 00 00',      # à adapter
        'address': 'N\'Djaména, Tchad',    # à adapter
    }

def jwt_required(view_func):
    """Vérifie manuellement le token JWT et attache request.user."""
    def wrapper(request, *args, **kwargs):
        auth = JWTAuthentication()
        auth_header = request.headers.get('Authorization', '')
        token = auth_header.split(' ')[-1] if auth_header.startswith('Bearer ') else ''
        try:
            validated_token = auth.get_validated_token(token)
            user = auth.get_user(validated_token)
            request.user = user
        except (InvalidToken, TokenError, IndexError):
            return JsonResponse({"detail": "Authentification requise."}, status=401)
        return view_func(request, *args, **kwargs)
    return wrapper


def get_format(request):
    fmt = request.GET.get('format', 'pdf').lower()
    if fmt not in ['pdf', 'excel']:
        fmt = 'pdf'
    return fmt


@jwt_required
def sales_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sales = Sale.objects.all()
    if start_date:
        sales = sales.filter(created_at__date__gte=start_date)
    if end_date:
        sales = sales.filter(created_at__date__lte=end_date)
    sales = sales.order_by('-created_at')
    fmt = get_format(request)

    if fmt == 'pdf':
        total = sales.aggregate(total=Coalesce(Sum('total_amount'), 0.0, output_field=DecimalField()))['total']
        context = {
            'sales': sales,
            'total': total,
            'start_date': start_date,
            'end_date': end_date,
            'pharmacy': get_pharmacy_info(request),

        }
        html_string = render_to_string('reports/sales_report.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_ventes.pdf"'
        return response
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventes"
        headers = ['ID', 'Date', 'Client', 'Pharmacien', 'Total', 'Mode de paiement']
        ws.append(headers)
        for sale in sales:
            ws.append([
                sale.id,
                sale.created_at.strftime('%d/%m/%Y %H:%M'),
                sale.customer_name or '',
                sale.user.email if sale.user else '',
                float(sale.total_amount),
                sale.get_payment_method_display()
            ])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rapport_ventes.xlsx"'
        return response


@jwt_required
def stock_report(request):
    lots = StockLot.objects.select_related('medicine').filter(expiry_date__gte=date.today(), quantity__gt=0).order_by('medicine__commercial_name')
    fmt = get_format(request)

    if fmt == 'pdf':
        total_value = lots.aggregate(total=Coalesce(Sum(F('quantity') * F('purchase_price_per_unit')), 0.0, output_field=DecimalField()))['total']
        context = {'lots': lots, 'total_value': total_value,
                   'pharmacy': get_pharmacy_info(request),}
        html_string = render_to_string('reports/stock_report.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_stocks.pdf"'
        return response
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stocks"
        headers = ['Médicament', 'Lot', 'Quantité', 'Date expiration', 'Prix unitaire']
        ws.append(headers)
        for lot in lots:
            ws.append([
                lot.medicine.commercial_name,
                lot.batch_number,
                lot.quantity,
                lot.expiry_date.strftime('%d/%m/%Y'),
                float(lot.purchase_price_per_unit)
            ])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rapport_stocks.xlsx"'
        return response


@jwt_required
def alert_report(request):
    notifs = Notification.objects.select_related('user').all().order_by('-created_at')
    fmt = get_format(request)

    if fmt == 'pdf':
        context = {'notifs': notifs,'pharmacy': get_pharmacy_info(request),}
        html_string = render_to_string('reports/alert_report.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_alertes.pdf"'
        return response
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Alertes"
        headers = ['Date', 'Utilisateur', 'Type', 'Message', 'Lu']
        ws.append(headers)
        for notif in notifs:
            ws.append([
                notif.created_at.strftime('%d/%m/%Y %H:%M'),
                notif.user.email,
                notif.get_type_display(),
                notif.message,
                'Oui' if notif.is_read else 'Non'
            ])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rapport_alertes.xlsx"'
        return response


@jwt_required
def user_report(request):
    users = User.objects.all()
    fmt = get_format(request)

    if fmt == 'pdf':
        context = {'users': users, 'pharmacy': get_pharmacy_info(request)}
        html_string = render_to_string('reports/user_report.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rapport_utilisateurs.pdf"'
        return response
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Utilisateurs"
        headers = ['Nom d\'utilisateur', 'Email', 'Rôle', 'Dernière connexion']
        ws.append(headers)
        for user in users:
            ws.append([
                user.username,
                user.email,
                user.get_role_display(),
                user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else ''
            ])
        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rapport_utilisateurs.xlsx"'
        return response